"""exports.py - data export helpers (CSV, XLSX, iCal, JSON) and planning rollups.

These take computed engine data (pass lists, equator crossings, Doppler
playbooks, comparisons) and serialize them to portable formats. Kept free of
Tk/GUI imports so they're unit-testable and reusable from reports or a CLI.
"""

import csv
import io
import json
import datetime as _dt


def _utc(unix, fmt="%Y-%m-%dT%H:%M:%SZ"):
    return _dt.datetime.fromtimestamp(
        unix, _dt.timezone.utc).strftime(fmt)


def _utc_human(unix, fmt="%Y-%m-%d %H:%M:%S UTC"):
    return _dt.datetime.fromtimestamp(
        unix, _dt.timezone.utc).strftime(fmt)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def rows_to_csv(headers, rows):
    """Return a CSV string from a header list and a list of row sequences."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    return buf.getvalue()


def passes_to_csv(passes, sat_name=""):
    """Serialize a list of PassPredict objects to CSV."""
    headers = ["satellite", "aos_utc", "los_utc", "tca_utc",
               "max_el_deg", "duration_min", "az_aos_deg", "az_los_deg"]
    rows = []
    for p in passes:
        dur = (p.los - p.aos) / 60.0 if p.los and p.aos else 0.0
        rows.append([
            sat_name, _utc_human(p.aos), _utc_human(p.los),
            _utc_human(p.tca), round(p.max_el, 1), round(dur, 1),
            round(p.az_aos, 1), round(p.az_los, 1),
        ])
    return rows_to_csv(headers, rows)


def crossings_to_csv(crossings, sat_name=""):
    """Serialize equator crossings [(unix, lon), ...] to CSV."""
    headers = ["satellite", "time_utc", "eqx_longitude_deg"]
    rows = [[sat_name, _utc_human(t), round(lon, 2)] for t, lon in crossings]
    return rows_to_csv(headers, rows)


def _hms(seconds):
    """Format a duration in seconds as H:MM:SS."""
    seconds = int(round(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    return "%d:%02d:%02d" % (h, m, s)


def eclipse_periods_rows(periods, sat_name=""):
    """Headers + rows for an orbit-by-orbit eclipse list. ``periods`` is a list
    of EclipsePeriod. Includes the interval (sunlit gap) between successive
    eclipses, matching Nova's Every-Orbit page."""
    headers = ["satellite", "enter_utc", "exit_utc", "duration",
               "interval_between", "sun_angle_deg"]
    out = []
    prev_exit = None
    for e in periods:
        interval = _hms(e.enter - prev_exit) if prev_exit is not None else ""
        out.append([
            sat_name, _utc_human(e.enter), _utc_human(e.exit),
            _hms(e.duration_s), interval, round(e.sun_angle, 1),
        ])
        prev_exit = e.exit
    return headers, out


def eclipse_periods_to_csv(periods, sat_name=""):
    headers, rows = eclipse_periods_rows(periods, sat_name)
    return rows_to_csv(headers, rows)


def eclipse_daily_rows(summary, sat_name=""):
    """Headers + rows for the daily eclipse summary. ``summary`` is a list of
    dicts from Predictor.eclipse_daily_summary."""
    headers = ["satellite", "date_utc", "eclipses", "total_eclipse",
               "longest_eclipse", "percent_of_day", "sun_angle_deg"]
    out = []
    for r in summary:
        out.append([
            sat_name, _utc_human(r["date"], "%Y-%m-%d"), r["count"],
            _hms(r["total_s"]), _hms(r["longest_s"]),
            round(r["percent"], 1), round(r["sun_angle"], 1),
        ])
    return headers, out


def eclipse_daily_to_csv(summary, sat_name=""):
    headers, rows = eclipse_daily_rows(summary, sat_name)
    return rows_to_csv(headers, rows)


def aoslos_rows(passes, sat_name=""):
    """Headers + rows for a compact AOS/LOS-only quick listing (many passes on
    one screen), matching Nova's AOS/LOS listing page."""
    headers = ["satellite", "aos_utc", "los_utc", "duration", "max_el_deg",
               "aos_az_deg", "los_az_deg"]
    out = []
    for p in passes:
        if not (p.aos and p.los):
            continue
        out.append([
            sat_name, _utc_human(p.aos), _utc_human(p.los),
            _hms(p.los - p.aos), round(p.max_el, 1),
            round(p.az_aos, 0), round(p.az_los, 0),
        ])
    return headers, out


def aoslos_to_csv(passes, sat_name=""):
    headers, rows = aoslos_rows(passes, sat_name)
    return rows_to_csv(headers, rows)


def listing_one_rows(rows, sat_name="", site_name=""):
    """Headers + rows for a one-observer stepped position listing."""
    headers = ["satellite", "site", "time_utc", "az_deg", "el_deg",
               "range_km", "range_rate_kms", "sub_lat", "sub_lon",
               "alt_km", "sunlit"]
    out = []
    for r in rows:
        out.append([
            sat_name, site_name, _utc_human(r["t"]),
            round(r["az"], 1), round(r["el"], 1), round(r["range_km"], 0),
            round(r["range_rate"], 3), round(r["sub_lat"], 2),
            round(r["sub_lon"], 2), round(r["alt_km"], 0),
            "yes" if r["sunlit"] else "no",
        ])
    return headers, out


def listing_one_to_csv(rows, sat_name="", site_name=""):
    headers, out = listing_one_rows(rows, sat_name, site_name)
    return rows_to_csv(headers, out)


def listing_two_rows(rows, sat_name="", site1="", site2=""):
    """Headers + rows for a two-observer stepped position listing (both stations'
    az/el/range from one ephemeris)."""
    headers = ["satellite", "time_utc",
               "%s_az" % (site1 or "obs1"), "%s_el" % (site1 or "obs1"),
               "%s_range_km" % (site1 or "obs1"),
               "%s_az" % (site2 or "obs2"), "%s_el" % (site2 or "obs2"),
               "%s_range_km" % (site2 or "obs2"),
               "sub_lat", "sub_lon", "alt_km"]
    out = []
    for r in rows:
        out.append([
            sat_name, _utc_human(r["t"]),
            round(r["az1"], 1), round(r["el1"], 1), round(r["range1_km"], 0),
            round(r["az2"], 1), round(r["el2"], 1), round(r["range2_km"], 0),
            round(r["sub_lat"], 2), round(r["sub_lon"], 2),
            round(r["alt_km"], 0),
        ])
    return headers, out


def listing_two_to_csv(rows, sat_name="", site1="", site2=""):
    headers, out = listing_two_rows(rows, sat_name, site1, site2)
    return rows_to_csv(headers, out)


def work_target_rows(windows, sat_name="", target=""):
    """Headers + rows for the Work-a-target shared-footprint windows. ``windows``
    are dicts from planning.best_passes_for_target."""
    headers = ["satellite", "target", "start_utc", "duration_min",
               "footprint_margin_deg"]
    out = []
    for w in windows:
        out.append([
            sat_name, target, _utc_human(w["start"]),
            round(w["duration_s"] / 60.0, 1), round(w["margin_deg"], 1),
        ])
    return headers, out


def visible_passes_rows(rows, sat_name=""):
    """Headers + rows for the optically-visible passes list. ``rows`` are dicts
    with aos/los/max_el/best_mag."""
    headers = ["satellite", "aos_utc", "los_utc", "duration_min",
               "max_el_deg", "est_magnitude"]
    out = []
    for r in rows:
        out.append([
            sat_name, _utc_human(r["aos"]), _utc_human(r["los"]),
            round((r["los"] - r["aos"]) / 60.0, 1),
            round(r["max_el"], 0), round(r["mag"], 1),
        ])
    return headers, out


def workable_rows(kind, items, sat_name="", when=""):
    """Headers + rows for the Workable list (grids / states / DXCC under the
    footprint). ``items`` is a flat list of strings."""
    headers = ["satellite", "kind", "when", kind]
    out = [[sat_name, kind, when, it] for it in items]
    return headers, out


def playbook_to_csv(rows, sat_name="", transponder=""):
    """Serialize Doppler playbook rows (from linkbudget.doppler_playbook_rows,
    with az/el attached by the Radio screen) to CSV. Frequencies in Hz."""
    headers = ["satellite", "transponder", "time_utc", "minute",
               "az_deg", "el_deg", "range_rate_kms", "rx_hz", "tx_hz", "mode"]
    out = []
    t0 = rows[0]["t"] if rows else 0
    for r in rows:
        out.append([
            sat_name, transponder, _utc_human(r["t"]),
            round((r["t"] - t0) / 60.0, 1),
            round(r.get("az", 0.0), 1), round(r.get("el", 0.0), 1),
            round(r["range_rate"], 4),
            r["rx_hz"], r["tx_hz"], r["mode"],
        ])
    return rows_to_csv(headers, out)


# ---------------------------------------------------------------------------
# XLSX (uses openpyxl if available; otherwise the caller should fall back to CSV)
# ---------------------------------------------------------------------------

def have_xlsx():
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def sheets_to_xlsx(path, sheets):
    """Write multiple sheets to an .xlsx file.

    ``sheets`` is a list of (title, headers, rows). Requires openpyxl; raises
    ImportError if unavailable so the caller can fall back to CSV.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0B3D91")
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title[:31] or "Sheet")
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
        for r in rows:
            ws.append(list(r))
        # autosize-ish: set a reasonable width from header length
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width = max(12, len(str(h)) + 2)
        ws.freeze_panes = "A2"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# iCalendar (.ics) pass schedule
# ---------------------------------------------------------------------------

def _ics_dt(unix):
    return _dt.datetime.fromtimestamp(
        unix, _dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def passes_to_ics(passes, sat_name="Satellite", site_name="", min_el=0.0,
                  prodid="-//OrbitDeck//Pass Schedule//EN"):
    """Build an iCalendar string with one VEVENT per pass (AOS->LOS), so passes
    show up in a calendar app with reminders. Times are UTC."""
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:" + prodid,
             "CALSCALE:GREGORIAN", "METHOD:PUBLISH"]
    for i, p in enumerate(passes):
        if not (p.aos and p.los):
            continue
        uid = "%s-%d-%s@orbitdeck" % (
            sat_name.replace(" ", "_"), i, _ics_dt(p.aos))
        summ = "%s pass (max el %.0f\u00b0)" % (sat_name, p.max_el)
        desc = ("Max elevation %.1f deg at %s. AOS az %.0f, LOS az %.0f. %s"
                % (p.max_el, _utc_human(p.tca), p.az_aos, p.az_los,
                   ("Station: " + site_name) if site_name else "")).strip()
        lines += [
            "BEGIN:VEVENT",
            "UID:" + uid,
            "DTSTAMP:" + _ics_dt(p.aos),
            "DTSTART:" + _ics_dt(p.aos),
            "DTEND:" + _ics_dt(p.los),
            "SUMMARY:" + summ,
            "DESCRIPTION:" + desc.replace("\n", " "),
            "BEGIN:VALARM",
            "TRIGGER:-PT10M",
            "ACTION:DISPLAY",
            "DESCRIPTION:" + summ,
            "END:VALARM",
            "END:VEVENT",
        ]
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


# ---------------------------------------------------------------------------
# JSON pass schedule
# ---------------------------------------------------------------------------

def passes_to_json(passes, sat_name="", site=None, min_el=0.0, indent=2):
    """Serialize passes to a JSON string with ISO-8601 UTC timestamps."""
    obj = {
        "satellite": sat_name,
        "site": site or {},
        "min_elevation_deg": min_el,
        "generated_utc": _utc(_now()),
        "passes": [],
    }
    for p in passes:
        if not (p.aos and p.los):
            continue
        obj["passes"].append({
            "aos_utc": _utc(p.aos),
            "los_utc": _utc(p.los),
            "tca_utc": _utc(p.tca),
            "max_elevation_deg": round(p.max_el, 2),
            "duration_s": round(p.los - p.aos, 1),
            "az_aos_deg": round(p.az_aos, 1),
            "az_los_deg": round(p.az_los, 1),
        })
    return json.dumps(obj, indent=indent)


def _now():
    import time
    return time.time()


# ---------------------------------------------------------------------------
# Multi-satellite comparison
# ---------------------------------------------------------------------------

def comparison_rows(entries):
    """Build a multi-satellite comparison table.

    ``entries`` is a list of dicts, each with: name, best_pass (PassPredict or
    None), n_passes, sunlit_frac (optional). Returns (headers, rows) suitable
    for CSV/XLSX/PDF.
    """
    headers = ["satellite", "passes_in_window", "best_pass_utc",
               "best_max_el_deg", "best_duration_min"]
    rows = []
    for e in entries:
        bp = e.get("best_pass")
        if bp and bp.aos:
            rows.append([
                e["name"], e.get("n_passes", 0), _utc_human(bp.aos),
                round(bp.max_el, 1),
                round((bp.los - bp.aos) / 60.0, 1) if bp.los else 0.0,
            ])
        else:
            rows.append([e["name"], e.get("n_passes", 0), "-", "-", "-"])
    return headers, rows


# ---------------------------------------------------------------------------
# Multi-site pass comparison
# ---------------------------------------------------------------------------

def site_comparison_rows(sat_name, site_entries):
    """Build a per-site comparison table for one satellite.

    ``site_entries``: list of dicts {name, n_passes, next_pass (PassPredict or
    None), best_pass (PassPredict or None)}.
    """
    headers = ["site", "satellite", "passes_in_window", "next_aos_utc",
               "next_max_el_deg", "best_max_el_deg"]
    rows = []
    for e in site_entries:
        nxt = e.get("next_pass")
        best = e.get("best_pass")
        rows.append([
            e["name"], sat_name, e.get("n_passes", 0),
            _utc_human(nxt.aos) if nxt and nxt.aos else "-",
            round(nxt.max_el, 1) if nxt and nxt.aos else "-",
            round(best.max_el, 1) if best and best.aos else "-",
        ])
    return headers, rows


# ---------------------------------------------------------------------------
# Satellite-to-satellite visibility windows
# ---------------------------------------------------------------------------

def sat2sat_rows(name1, name2, windows):
    headers = ["sat_a", "sat_b", "start_utc", "end_utc", "duration_min",
               "min_range_km"]
    rows = []
    for w in windows:
        rows.append([
            name1, name2, _utc_human(w["start"]), _utc_human(w["end"]),
            round(w.get("duration_s", 0) / 60.0, 1),
            round(w.get("min_range_km", 0), 0),
        ])
    return headers, rows


# ---------------------------------------------------------------------------
# Celestial body positions (snapshot)
# ---------------------------------------------------------------------------

def celestial_rows(entries):
    """entries: list of (name, az, el, extra). Returns (headers, rows)."""
    headers = ["body", "azimuth_deg", "elevation_deg", "notes"]
    rows = []
    for name, az, el, extra in entries:
        rows.append([name, round(az, 1), round(el, 1), extra])
    return headers, rows


# ---------------------------------------------------------------------------
# EME common-Moon windows
# ---------------------------------------------------------------------------

def eme_window_rows(site_a, site_b, windows):
    headers = ["station_a", "station_b", "start_utc", "end_utc",
               "duration_min"]
    rows = []
    for (start, end) in windows:
        rows.append([site_a, site_b, _utc_human(start), _utc_human(end),
                     round((end - start) / 60.0, 1)])
    return headers, rows
